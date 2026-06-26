"""Analítica del comparativo: mapeo 1 a 1, outliers y comparativo regional."""
from __future__ import annotations

import re
import io
from typing import Optional

import numpy as np
import pandas as pd

from .nlp_mapper import normalize


def _cop(x) -> str:
    """Formatea un número como pesos colombianos: 19800 -> '$19.800'."""
    try:
        if x is None:
            return "—"
        return "$" + f"{float(x):,.0f}".replace(",", ".")
    except Exception:
        return str(x)


# ---------------------------------------------------------------------------
# 1. Mapeo 1 a 1 (ya construido en el pipeline; aquí solo se formatea)
# ---------------------------------------------------------------------------

def build_mapping_report(matches: pd.DataFrame) -> pd.DataFrame:
    """Detalle del cruce warehouse <-> fuente, con la fuente que refuta el
    precio, el enlace a ella y la diferencia frente al escenario solo-IPC."""
    cols = [
        "codigo", "descripcion_wh", "und_wh", "grupo", "clasificacion", "categoria",
        "candidato", "score", "confianza_match", "metodo", "unidad_coincide",
        "valor_wh", "valor_wh_proyectado", "factor_proyeccion",
        "precio_comparativo_promedio", "precio_comparativo_min",
        "precio_comparativo_mediana", "precio_comparativo_max", "n_cotizaciones",
        "region_mejor_comparativo", "proveedor_comparativo",
        "region_comparativo_max", "proveedor_comparativo_max", "archivo_comparativo_max",
        "todas_las_fuentes",
        "precio_consolidado_promedio", "precio_consolidado_mediana",
        "precio_consolidado_min", "precio_consolidado_max", "n_facturas_consolidado",
        "consumo_anual",
        "precio_referencia", "como_se_calculo", "de_donde_salio_el_precio",
        "fuente_que_refuta", "enlace_fuente",
        "diferencia_vs_ipc", "pct_diferencia", "sospechoso_dif_mayor_50pct",
        "warehouse_por_debajo_del_mercado",
        "descartado_por_magnitud",
        "consumo_usado", "ahorro_ponderado", "anio_actualizado",
        "nuevo_valor", "actualizado",
    ]
    existing = [c for c in cols if c in matches.columns]
    out = matches[existing].copy()
    if out.empty or "score" not in out.columns:
        return out
    return out.sort_values("score", ascending=False).reset_index(drop=True)


# ---------------------------------------------------------------------------
# 2. Análisis de outliers por ítem (IQR + Z-Score)
# ---------------------------------------------------------------------------

def outlier_analysis(comparativos: pd.DataFrame, z_thresh: float = 3.0) -> pd.DataFrame:
    """Identifica precios atípicos por ítem usando IQR y Z-Score.

    Agrupa por ítem normalizado (todas las regiones/proveedores juntos) y marca
    cada observación como outlier por IQR (1.5*IQR) y/o por |z| > z_thresh.
    """
    if comparativos.empty:
        return pd.DataFrame()

    df = comparativos.copy()
    df["item_norm"] = df["descripcion"].map(normalize)
    df = df[df["precio"].notna() & df["item_norm"].ne("")]

    rows = []
    for item, g in df.groupby("item_norm"):
        precios = g["precio"].astype(float)
        if len(precios) < 3:
            # Sin suficientes observaciones para estadística robusta.
            continue
        q1, q3 = precios.quantile(0.25), precios.quantile(0.75)
        iqr = q3 - q1
        low, high = q1 - 1.5 * iqr, q3 + 1.5 * iqr
        mean, std = precios.mean(), precios.std(ddof=0)
        for idx, p in precios.items():
            z = (p - mean) / std if std and std > 0 else 0.0
            is_iqr = bool(p < low or p > high)
            is_z = bool(abs(z) > z_thresh)
            if is_iqr or is_z:
                rows.append(
                    {
                        "item_norm": item,
                        "descripcion": g.loc[idx, "descripcion"],
                        "region": g.loc[idx, "region"],
                        "proveedor": g.loc[idx, "proveedor"],
                        "precio": round(float(p), 2),
                        "mediana_item": round(float(precios.median()), 2),
                        "q1": round(float(q1), 2),
                        "q3": round(float(q3), 2),
                        "iqr": round(float(iqr), 2),
                        "limite_inf": round(float(low), 2),
                        "limite_sup": round(float(high), 2),
                        "z_score": round(float(z), 2),
                        "outlier_iqr": is_iqr,
                        "outlier_zscore": is_z,
                        "n_observaciones": int(len(precios)),
                    }
                )
    if not rows:
        return pd.DataFrame()
    return pd.DataFrame(rows).sort_values(["item_norm", "precio"]).reset_index(drop=True)


# ---------------------------------------------------------------------------
# 3. Comparativo regional (dispersión y mediana por ítem y región)
# ---------------------------------------------------------------------------

def regional_comparison(comparativos: pd.DataFrame) -> pd.DataFrame:
    """Tabla larga: por ítem y región -> mediana, min, max, std, n."""
    if comparativos.empty:
        return pd.DataFrame()
    df = comparativos.copy()
    df["item_norm"] = df["descripcion"].map(normalize)
    df = df[df["precio"].notna() & df["item_norm"].ne("")]

    agg = (
        df.groupby(["item_norm", "region"])["precio"]
        .agg(["count", "min", "median", "max", "mean", "std"])
        .reset_index()
        .rename(
            columns={
                "count": "n",
                "min": "precio_min",
                "median": "precio_mediana",
                "max": "precio_max",
                "mean": "precio_promedio",
                "std": "desv_std",
            }
        )
    )
    # Una descripción representativa por ítem (la más frecuente).
    repr_desc = (
        df.groupby("item_norm")["descripcion"]
        .agg(lambda s: s.mode().iat[0] if not s.mode().empty else s.iloc[0])
        .reset_index()
        .rename(columns={"descripcion": "descripcion"})
    )
    agg = agg.merge(repr_desc, on="item_norm", how="left")
    for c in ["precio_min", "precio_mediana", "precio_max", "precio_promedio", "desv_std"]:
        agg[c] = agg[c].round(2)
    return agg.sort_values(["item_norm", "region"]).reset_index(drop=True)


def regional_pivot(comparativos: pd.DataFrame) -> pd.DataFrame:
    """Tabla pivote: filas=ítem, columnas=región, valores=mediana de precio.

    Incluye una columna de dispersión inter-regional (coef. de variación).
    """
    reg = regional_comparison(comparativos)
    if reg.empty:
        return pd.DataFrame()
    pivot = reg.pivot_table(
        index="item_norm", columns="region", values="precio_mediana", aggfunc="first"
    )
    desc = reg.drop_duplicates("item_norm").set_index("item_norm")["descripcion"]
    pivot.insert(0, "descripcion", desc)
    # Dispersión inter-regional sobre las medianas regionales.
    region_cols = [c for c in pivot.columns if c != "descripcion"]
    medianas = pivot[region_cols]
    pivot["mediana_global"] = medianas.median(axis=1).round(2)
    pivot["dispersion_cv_%"] = (
        (medianas.std(axis=1) / medianas.mean(axis=1) * 100).round(1)
    )
    return pivot.reset_index()


# ---------------------------------------------------------------------------
# Exportación del reporte (xlsx en memoria)
# ---------------------------------------------------------------------------

def _clasificacion_frame(stats: dict) -> pd.DataFrame:
    """Resumen de la columna Clasificación de PRIMARIOS: cuántas actividades hay
    de cada estado en el warehouse y, de las Activas, cuántas son actualizables."""
    cl = (stats or {}).get("clasificacion", {})
    dist = cl.get("distribucion", {}) or {}
    valor_activa = str(cl.get("valor_activa", "Activa"))
    filas = []
    total = 0
    # Orden: Activa primero, luego el resto por cantidad descendente.
    items = sorted(dist.items(), key=lambda kv: (kv[0].strip().lower() != valor_activa.strip().lower(), -kv[1]))
    for nombre, cant in items:
        filas.append({"Clasificación": nombre, "Cantidad": int(cant)})
        total += int(cant)
    if filas:
        filas.append({"Clasificación": "TOTAL", "Cantidad": total})
    evaluables = cl.get("activas_evaluables")
    actualizables = cl.get("activas_actualizables")
    if evaluables is not None:
        filas.append({
            "Clasificación": f"{valor_activa} evaluables (grupo cruzable + precio)",
            "Cantidad": int(evaluables),
        })
    if actualizables is not None:
        filas.append({
            "Clasificación": f"{valor_activa} actualizables (con fuente que refuta)",
            "Cantidad": int(actualizables),
        })
    return pd.DataFrame(filas) if filas else pd.DataFrame({"info": ["sin datos de clasificación"]})


def _categoria_frame(stats: dict) -> pd.DataFrame:
    pc = (stats or {}).get("por_categoria", {})
    filas = []
    for cat in ("Material", "Mano de obra", "Viáticos"):
        d = pc.get(cat, {})
        filas.append({
            "Categoría": cat,
            "Ítems cruzados": d.get("cruces", 0),
            "Más barato que IPC": d.get("mas_barato_que_ipc", 0),
            "Más caro que IPC": d.get("mas_caro_que_ipc", 0),
            "Ahorro potencial (COP)": d.get("ahorro_potencial", 0),
            "Sobrecosto potencial (COP)": d.get("sobrecosto_potencial", 0),
            "Diferencia neta (COP)": d.get("diferencia_neta", 0),
        })
    return pd.DataFrame(filas)


def _parse_fuentes(texto: str):
    """Extrae (etiqueta, precio) de la cadena todas_las_fuentes."""
    out = []
    if not texto or not isinstance(texto, str):
        return out
    for part in texto.split(" ; "):
        part = part.strip()
        if not part or part.startswith("(+"):
            continue
        if "$" in part:
            etiqueta, _, pr = part.rpartition("$")
            etiqueta = etiqueta.rstrip(": ").strip()
            num = re.sub(r"[^\d]", "", pr.split(",")[0])
            try:
                precio = float(num) if num else None
            except Exception:
                precio = None
            out.append((etiqueta or pr, precio))
    return out


def build_items_revisar_sheet(writer, mapping: pd.DataFrame, top_impacto: int = 15):
    """Crea la hoja 'Items a Revisar' con formato legible: un bloque por ítem con
    el valor de la BD, mínimo y máximo, y la lista de registros (extremos
    subrayados). Incluye ítems con (a) registros muy dispersos entre sí, (b) todos
    los valores muy distintos a la BD, y (c) mayor impacto en ahorro/sobrecosto."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    book = writer.book
    ws = book.create_sheet("Items a Revisar")
    ws.sheet_properties.tabColor = "E8742C"

    NARANJA = "E8742C"; GRIS = "F2F2F2"; ROJO = "C0392B"; VERDE_OSC = "1E6B3A"
    titulo_f = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    item_f = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    razon_f = Font(name="Calibri", size=10, italic=True, color="FFF2E6")
    lbl_f = Font(name="Calibri", size=10, bold=True, color="333333")
    val_f = Font(name="Calibri", size=10, color="333333")
    rec_f = Font(name="Calibri", size=10, color="444444")
    ext_f = Font(name="Calibri", size=10, bold=True, underline="single", color=ROJO)
    naranja_fill = PatternFill("solid", fgColor=NARANJA)
    gris_fill = PatternFill("solid", fgColor=GRIS)

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22

    ws.merge_cells("A1:D1")
    t = ws["A1"]; t.value = "ITEMS A REVISAR — valores dispersos o muy distintos a la BD"
    t.font = titulo_f; t.fill = naranja_fill; t.alignment = Alignment("left", "center")
    ws.row_dimensions[1].height = 26
    r = 3

    if mapping is None or mapping.empty:
        ws.cell(r, 2, "Sin ítems para revisar.").font = val_f
        return

    base_col = "valor_wh_proyectado" if "valor_wh_proyectado" in mapping.columns else "valor_wh"
    df = mapping.copy()
    df["_bd"] = pd.to_numeric(df.get(base_col), errors="coerce")
    df["_ap"] = pd.to_numeric(df.get("ahorro_ponderado"), errors="coerce").fillna(0)
    df["_sc"] = pd.to_numeric(df.get("score"), errors="coerce").fillna(0)

    SCORE_MIN = 95.0   # el match debe ser de al menos 95%
    DESV = 0.70        # solo registros que difieren más de ±70% de la BD
    seleccion = []     # (orden, impacto, row, razon, registros, vmin, vmax, bd)

    for idx, row in df.iterrows():
        if row["_sc"] < SCORE_MIN:
            continue
        recs = _parse_fuentes(row.get("todas_las_fuentes"))
        precios = [p for _, p in recs if p and p > 0]
        bd = row["_bd"]
        if len(precios) < 1 or not bd or bd <= 0:
            continue
        vmin, vmax = min(precios), max(precios)
        # Gate: al menos un registro difiere más de ±70% de la BD.
        fuera_70 = (vmin < bd * (1 - DESV)) or (vmax > bd * (1 + DESV))
        if not fuera_70:
            continue
        disp = (vmax / vmin) if vmin > 0 else 1
        if disp >= 3 and len(precios) >= 2:
            razon = "Tiene registros con valores muy diferentes entre sí."; orden = 0
        elif (vmin > bd * (1 + DESV)) or (vmax < bd * (1 - DESV)):
            razon = "Todos los registros difieren notablemente del valor de la BD."; orden = 1
        else:
            razon = "Tiene registros que difieren más del 70% del valor de la BD."; orden = 2
        seleccion.append((orden, abs(row["_ap"]), row, razon, recs, vmin, vmax, bd))

    seleccion.sort(key=lambda x: (x[0], -x[1]))
    if not seleccion:
        ws.cell(r, 2, "Sin ítems con match ≥95% y diferencias mayores al 70%.").font = val_f
        return

    thin = Side(style="thin", color="DDDDDD")
    for _, _, row, razon, recs, vmin, vmax, bd in seleccion:
        # Encabezado del ítem
        ws.merge_cells(f"A{r}:D{r}")
        ce = ws.cell(r, 1, f"  {row.get('descripcion_wh','')}")
        ce.font = item_f; ce.fill = naranja_fill; ce.alignment = Alignment("left", "center")
        ws.row_dimensions[r].height = 20
        r += 1
        ws.merge_cells(f"A{r}:D{r}")
        cr = ws.cell(r, 1, f"  {razon}")
        cr.font = razon_f; cr.fill = naranja_fill
        r += 1
        clasif = str(row.get("clasificacion") or "").strip()
        if clasif and clasif.lower() not in ("nan", "none"):
            ws.cell(r, 2, "Clasificación:").font = lbl_f; ws.cell(r, 2).alignment = Alignment("right")
            ws.cell(r, 3, clasif).font = val_f
            r += 1
        # Valores BD / min / max
        ws.cell(r, 2, "Valor de la BD (Wh):").font = lbl_f
        ws.cell(r, 2).alignment = Alignment("right")
        ws.cell(r, 3, _cop(bd) if bd else "—").font = Font(bold=True, color=VERDE_OSC)
        r += 1
        ws.cell(r, 2, "Valor mínimo:").font = lbl_f; ws.cell(r, 2).alignment = Alignment("right")
        ws.cell(r, 3, _cop(vmin)).font = val_f
        ws.cell(r, 2).fill = gris_fill; ws.cell(r, 3).fill = gris_fill
        r += 1
        ws.cell(r, 2, "Valor máximo:").font = lbl_f; ws.cell(r, 2).alignment = Alignment("right")
        ws.cell(r, 3, _cop(vmax)).font = val_f
        r += 1
        ws.cell(r, 2, "Registros (extremos subrayados en rojo):").font = lbl_f
        r += 1
        # Lista de registros, extremos resaltados
        recs_sorted = sorted([x for x in recs if x[1]], key=lambda x: x[1])
        for etiqueta, precio in recs_sorted[:40]:
            es_extremo = precio in (vmin, vmax)
            ws.cell(r, 2, f"   • {etiqueta}")
            ws.cell(r, 2).font = ext_f if es_extremo else rec_f
            ws.cell(r, 3, _cop(precio)).font = ext_f if es_extremo else rec_f
            for col in (2, 3):
                ws.cell(r, col).border = Border(bottom=thin)
            r += 1
        if len(recs_sorted) > 40:
            ws.cell(r, 2, f"   (+{len(recs_sorted) - 40} registros más)").font = razon_f
            ws.cell(r, 2).font = Font(italic=True, color="999999")
            r += 1
        r += 1  # separador


def build_excel_report(
    mapping: pd.DataFrame,
    outliers: pd.DataFrame,
    regional: pd.DataFrame,
    regional_pivot_df: pd.DataFrame,
    stats: Optional[dict] = None,
    conclusiones: Optional[list] = None,
    consolidado_planta: Optional[pd.DataFrame] = None,
) -> bytes:
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # Resumen ejecutivo primero.
        if stats:
            _stats_to_frame(stats, conclusiones).to_excel(
                writer, sheet_name="Resumen Ejecutivo", index=False
            )
            _categoria_frame(stats).to_excel(
                writer, sheet_name="Resumen por Categoria", index=False
            )
            _clasificacion_frame(stats).to_excel(
                writer, sheet_name="Resumen por Clasificacion", index=False
            )
        (mapping if not mapping.empty else pd.DataFrame({"info": ["sin datos"]})).to_excel(
            writer, sheet_name="Mapping y Refutacion", index=False
        )
        if consolidado_planta is not None and not consolidado_planta.empty:
            consolidado_planta.sort_values(["insumo", "planta_region"]).to_excel(
                writer, sheet_name="Consolidado por Planta", index=False
            )
        (outliers if not outliers.empty else pd.DataFrame({"info": ["sin outliers"]})).to_excel(
            writer, sheet_name="Analisis Outliers", index=False
        )
        (regional if not regional.empty else pd.DataFrame({"info": ["sin datos"]})).to_excel(
            writer, sheet_name="Comparativo Regional", index=False
        )
        _format_sheets(writer)
        # Hoja 'Items a Revisar' con formato legible (bloques por ítem). Se crea
        # después de _format_sheets para que su formato manual no se altere.
        build_items_revisar_sheet(writer, mapping)
    buffer.seek(0)
    return buffer.getvalue()


def _items_a_revisar(mapping: pd.DataFrame, umbral: float = 1.5) -> pd.DataFrame:
    """Selecciona los cruces cuya referencia difiere del valor de la BD por más
    de ±50% (ref > 1.5x BD o < 0.5x BD), para auditoría manual. Ordena por la
    desviación (los más alejados primero)."""
    if mapping is None or mapping.empty:
        return pd.DataFrame()
    df = mapping.copy()
    base_col = "valor_wh_proyectado" if "valor_wh_proyectado" in df.columns else "valor_wh"
    if "precio_referencia" not in df.columns or base_col not in df.columns:
        return pd.DataFrame()
    df = df[pd.to_numeric(df["precio_referencia"], errors="coerce").notna()]
    df = df[pd.to_numeric(df[base_col], errors="coerce") > 0]
    if df.empty:
        return pd.DataFrame()
    ratio = pd.to_numeric(df["precio_referencia"], errors="coerce") / pd.to_numeric(df[base_col], errors="coerce")
    df = df.assign(
        relacion_ref_vs_bd=ratio.round(2),
        pct_vs_bd=((ratio - 1.0) * 100).round(1),  # + = mercado más caro que la BD
    )
    df = df[(ratio > 1.5) | (ratio < 0.5)]   # más de ±50%
    if df.empty:
        return pd.DataFrame()
    df = df.assign(_dev=(ratio[df.index] - 1.0).abs()).sort_values("_dev", ascending=False)
    cols = [c for c in [
        "codigo", "descripcion_wh", "und_wh", "categoria", "score", "candidato",
        base_col, "precio_referencia", "relacion_ref_vs_bd", "pct_vs_bd",
        "precio_consolidado_promedio", "n_facturas_consolidado",
        "precio_comparativo_promedio", "precio_comparativo_max", "n_cotizaciones",
        "como_se_calculo", "de_donde_salio_el_precio", "todas_las_fuentes",
    ] if c in df.columns]
    return df[cols].reset_index(drop=True)


def _format_sheets(writer) -> None:
    """Ajusta el ancho de cada columna para que el título completo sea visible,
    congela la fila de encabezado y activa el autofiltro."""
    from openpyxl.utils import get_column_letter
    for ws in writer.book.worksheets:
        max_row = ws.max_row
        for col_idx in range(1, ws.max_column + 1):
            letter = get_column_letter(col_idx)
            header = ws.cell(row=1, column=col_idx).value
            best = len(str(header)) if header is not None else 10
            # Muestra hasta 80 filas para estimar el ancho del contenido.
            for r in range(2, min(max_row, 80) + 1):
                v = ws.cell(row=r, column=col_idx).value
                if v is not None:
                    best = max(best, len(str(v)))
            ws.column_dimensions[letter].width = min(max(best + 2, 12), 60)
        if max_row >= 1 and ws.max_column >= 1:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{max_row}"


# ---------------------------------------------------------------------------
# Estadísticos y conclusiones
# ---------------------------------------------------------------------------

def _region_cost_index(comparativos: pd.DataFrame) -> pd.DataFrame:
    """Índice de costo por región: para cada ítem presente en varias regiones,
    se calcula precio_region / mediana_global_item; el promedio por región da un
    índice >1 (más caro que el promedio) o <1 (más económico)."""
    if comparativos.empty:
        return pd.DataFrame()
    df = comparativos.copy()
    df["item_norm"] = df["descripcion"].map(normalize)
    df = df[df["precio"].notna() & df["item_norm"].ne("")]
    med_region = df.groupby(["item_norm", "region"])["precio"].median().reset_index()
    med_global = df.groupby("item_norm")["precio"].median().rename("med_global").reset_index()
    merged = med_region.merge(med_global, on="item_norm")
    merged = merged[merged["med_global"] > 0]
    merged["ratio"] = merged["precio"] / merged["med_global"]
    idx = (
        merged.groupby("region")
        .agg(indice_costo=("ratio", "mean"), items=("item_norm", "nunique"))
        .reset_index()
    )
    # Solo regiones con respaldo suficiente.
    idx = idx[idx["items"] >= 3]
    idx["indice_costo"] = idx["indice_costo"].round(3)
    return idx.sort_values("indice_costo", ascending=False).reset_index(drop=True)


def compute_stats(matches: pd.DataFrame, comparativos: pd.DataFrame) -> dict:
    s: dict = {}
    # Fuentes analizadas.
    if not comparativos.empty and "archivo" in comparativos.columns:
        s["fuentes_analizadas"] = int(comparativos["archivo"].nunique())
        s["fuentes_listado"] = sorted(comparativos["archivo"].dropna().unique().tolist())
    else:
        s["fuentes_analizadas"] = 0
        s["fuentes_listado"] = []
    s["regiones_analizadas"] = (
        sorted(comparativos["region"].dropna().unique().tolist())
        if not comparativos.empty else []
    )
    s["registros_precio"] = int(len(comparativos))

    # Top regiones más caras / económicas.
    idx = _region_cost_index(comparativos)
    if not idx.empty:
        s["top5_regiones_mas_costosas"] = idx.head(5)[["region", "indice_costo"]].values.tolist()
        s["top5_regiones_mas_economicas"] = (
            idx.tail(5).sort_values("indice_costo")[["region", "indice_costo"]].values.tolist()
        )
    else:
        s["top5_regiones_mas_costosas"] = []
        s["top5_regiones_mas_economicas"] = []

    # Precios por debajo / por encima del escenario solo-IPC.
    if not matches.empty and "diferencia_vs_ipc" in matches.columns:
        m = matches[matches["actualizado"] == True].copy()  # noqa: E712
        m = m[m["diferencia_vs_ipc"].notna()]
        s["cruces_validos"] = int(len(m))
        # Columna de ahorro: ponderada por consumo si existe; si no, por unidad.
        val_col = "ahorro_ponderado" if "ahorro_ponderado" in m.columns else "diferencia_vs_ipc"
        m[val_col] = pd.to_numeric(m[val_col], errors="coerce")
        m = m[m[val_col].notna()]
        # mercado por DEBAJO del IPC = se consigue más barato (diferencia > 0).
        below = m[m[val_col] > 0]
        above = m[m[val_col] < 0]
        s["items_mercado_mas_barato_que_ipc"] = int(len(below))
        s["items_mercado_mas_caro_que_ipc"] = int(len(above))
        s["ahorro_potencial_total"] = round(float(below[val_col].sum()), 0)
        s["sobrecosto_potencial_total"] = round(float(-above[val_col].sum()) + 0.0, 0)
        s["diferencia_neta_total"] = round(float(m[val_col].sum()), 0)
        s["ahorro_ponderado_por_consumo"] = bool(val_col == "ahorro_ponderado")
        s["items_sospechosos_mayor_50pct"] = int(
            matches["sospechoso_dif_mayor_50pct"].fillna(False).sum()
        ) if "sospechoso_dif_mayor_50pct" in matches.columns else 0
        # Base anual (valor BD proyectado x consumo) para expresar el ahorro en %.
        base_col_v = "valor_wh_proyectado" if "valor_wh_proyectado" in m.columns else "valor_wh"
        if base_col_v in m.columns:
            qcol = "consumo_usado" if "consumo_usado" in m.columns else None
            base = pd.to_numeric(m[base_col_v], errors="coerce").fillna(0)
            if qcol:
                base = base * pd.to_numeric(m[qcol], errors="coerce").fillna(1)
            total_base = float(base.sum())
            if total_base > 0:
                s["base_anual_total"] = round(total_base, 0)
                s["ahorro_potencial_pct"] = round(s["ahorro_potencial_total"] / total_base * 100, 1)
                s["sobrecosto_potencial_pct"] = round(s["sobrecosto_potencial_total"] / total_base * 100, 1)
                s["diferencia_neta_pct"] = round(s["diferencia_neta_total"] / total_base * 100, 1)

        # Segmentación por categoría: Material | Mano de obra | Viáticos.
        por_cat = {}
        cat_col = "categoria" if "categoria" in m.columns else None
        cats = ["Material", "Mano de obra", "Viáticos"]
        for cat in cats:
            sub = m[m[cat_col] == cat] if cat_col else m.iloc[0:0]
            below_c = sub[sub[val_col] > 0]
            above_c = sub[sub[val_col] < 0]
            por_cat[cat] = {
                "cruces": int(len(sub)),
                "mas_barato_que_ipc": int(len(below_c)),
                "mas_caro_que_ipc": int(len(above_c)),
                "ahorro_potencial": round(float(below_c[val_col].sum()), 0),
                "sobrecosto_potencial": round(float(-above_c[val_col].sum()) + 0.0, 0),
                "diferencia_neta": round(float(sub[val_col].sum()), 0),
            }
        s["por_categoria"] = por_cat
    else:
        s.update({
            "cruces_validos": 0,
            "items_mercado_mas_barato_que_ipc": 0,
            "items_mercado_mas_caro_que_ipc": 0,
            "ahorro_potencial_total": 0,
            "sobrecosto_potencial_total": 0,
            "diferencia_neta_total": 0,
        })
        s["por_categoria"] = {}
    return s


def build_conclusions(s: dict) -> list:
    def cop(x):
        try:
            return "$" + f"{float(x):,.0f}".replace(",", ".")
        except Exception:
            return str(x)

    c = []
    c.append(
        f"Se analizaron {s.get('fuentes_analizadas', 0)} fuentes de datos "
        f"({s.get('registros_precio', 0)} precios) en "
        f"{len(s.get('regiones_analizadas', []))} regiones."
    )
    c.append(
        f"De {s.get('cruces_validos', 0)} ítems cruzados, en "
        f"{s.get('items_mercado_mas_barato_que_ipc', 0)} el mercado ofrece un precio MENOR "
        f"al que daría aplicar solo el IPC (oportunidad de ahorro), y en "
        f"{s.get('items_mercado_mas_caro_que_ipc', 0)} el mercado está por ENCIMA del IPC."
    )
    def _pct(k):
        v = s.get(k)
        return f" ({v:.1f}% del gasto base)" if v is not None else ""
    c.append(
        f"Ahorro potencial: {cop(s.get('ahorro_potencial_total', 0))}{_pct('ahorro_potencial_pct')}. "
        f"Sobrecosto potencial: {cop(s.get('sobrecosto_potencial_total', 0))}{_pct('sobrecosto_potencial_pct')}. "
        f"Diferencia neta: {cop(s.get('diferencia_neta_total', 0))}{_pct('diferencia_neta_pct')} "
        "(positivo = el mercado está, en neto, por debajo del ajuste proyectado)."
    )
    if s.get("items_sospechosos_mayor_50pct"):
        c.append(
            f"{s['items_sospechosos_mayor_50pct']} ítems tienen una diferencia mayor al 50% "
            "frente a la BD: se marcaron como SOSPECHOSOS, no se actualizan al valor de "
            "mercado y quedan en la hoja 'Items a Revisar' para decisión manual."
        )
    if s.get("top5_regiones_mas_costosas"):
        caras = ", ".join(f"{r} ({i:.2f})" for r, i in s["top5_regiones_mas_costosas"][:3])
        c.append(f"Regiones más costosas (índice vs. mediana nacional): {caras}.")
    if s.get("top5_regiones_mas_economicas"):
        baratas = ", ".join(f"{r} ({i:.2f})" for r, i in s["top5_regiones_mas_economicas"][:3])
        c.append(f"Regiones más económicas: {baratas}.")
    pc = s.get("por_categoria", {})
    for cat in ("Material", "Mano de obra", "Viáticos"):
        if cat in pc and pc[cat]["cruces"] > 0:
            d = pc[cat]
            c.append(
                f"[{cat}] {d['cruces']} ítems · ahorro {cop(d['ahorro_potencial'])} · "
                f"sobrecosto {cop(d['sobrecosto_potencial'])} · neta {cop(d['diferencia_neta'])}."
            )
    cl = s.get("clasificacion", {})
    if cl.get("activas_evaluables") is not None:
        va = str(cl.get("valor_activa", "Activa"))
        ev = int(cl.get("activas_evaluables", 0))
        act = int(cl.get("activas_actualizables", 0))
        pct = f" ({act/ev*100:.1f}%)" if ev else ""
        c.append(
            f"De {ev} actividades {va} evaluables, es posible actualizar {act}{pct} "
            f"(las demás se mantienen en la BD por no tener fuente que las refute o "
            f"por diferencia sospechosa). Ver hoja 'Resumen por Clasificacion'."
        )
    return c


def _stats_to_frame(s: dict, conclusiones: Optional[list]) -> pd.DataFrame:
    filas = [
        ("Fuentes de datos analizadas", s.get("fuentes_analizadas", 0)),
        ("Registros de precio", s.get("registros_precio", 0)),
        ("Regiones analizadas", len(s.get("regiones_analizadas", []))),
        ("Ítems cruzados (válidos)", s.get("cruces_validos", 0)),
        ("Ítems con mercado más barato que IPC", s.get("items_mercado_mas_barato_que_ipc", 0)),
        ("Ítems con mercado más caro que IPC", s.get("items_mercado_mas_caro_que_ipc", 0)),
        ("Ahorro potencial total (COP)", s.get("ahorro_potencial_total", 0)),
        ("Sobrecosto potencial total (COP)", s.get("sobrecosto_potencial_total", 0)),
        ("Diferencia neta total (COP)", s.get("diferencia_neta_total", 0)),
    ]
    for i, (r, v) in enumerate(s.get("top5_regiones_mas_costosas", []), 1):
        filas.append((f"Región más costosa #{i}", f"{r} (índice {v:.2f})"))
    for i, (r, v) in enumerate(s.get("top5_regiones_mas_economicas", []), 1):
        filas.append((f"Región más económica #{i}", f"{r} (índice {v:.2f})"))
    cl = s.get("clasificacion", {})
    dist = cl.get("distribucion", {}) or {}
    for nombre, cant in sorted(dist.items(), key=lambda kv: -kv[1]):
        filas.append((f"Clasificación · {nombre}", int(cant)))
    if cl.get("activas_evaluables") is not None:
        va = str(cl.get("valor_activa", "Activa"))
        filas.append((f"{va} evaluables", int(cl.get("activas_evaluables", 0))))
        filas.append((f"{va} actualizables", int(cl.get("activas_actualizables", 0))))
    for i, txt in enumerate(conclusiones or [], 1):
        filas.append((f"Conclusión {i}", txt))
    return pd.DataFrame(filas, columns=["Métrica", "Valor"])
